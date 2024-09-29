import pandas as pd
from openpyxl import Workbook

# 生成dprate和dropout的取值范围
dprate_range = dropout_range = [0.1 ,0.2 ,0.3 ,0.4 ,0.5 ,0.6 ,0.7 ,0.8 ]

# 生成空的DataFrame
df = pd.DataFrame(index=dprate_range, columns=dropout_range)
K=5
# for K in range(1,11):
# 从日志文件中读取数据并填充DataFrame
with open('result22/K_{}.txt'.format(K), 'r') as f:
    i=0
    for line in f:
        if line.startswith('INFO:root:Namespace(K={}'.format(K)):
            params = line.strip().split(',')
            dprate = float(params[6].split('=')[1])
            dropout = float(params[7].split('=')[1])
        elif line.startswith('INFO:root:0'):
            score = float(line.strip().split(':')[2])
            i=i+1
            if i%2!=0:
                df.at[dprate, dropout] = round(score,4)

# 创建Excel文件并将DataFrame写入其中
wb = Workbook()
ws = wb.active
ws.title = 'K_{}'.format(K)
for i, dprate in enumerate(dprate_range):
    for j, dropout in enumerate(dropout_range):
        ws.cell(row=i+2, column=j+2, value=df.at[dprate, dropout])
        if i == 0:
            ws.cell(row=1, column=j+2, value=f'Dropout={dropout}')
        if j == 0:
            ws.cell(row=i+2, column=1, value=f'Dprate={dprate}')

# 保存Excel文件
wb.save('excel/acmv9_K_{}.xlsx'.format(K))